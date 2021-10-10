import pandas as pd
from openpyxl import load_workbook
import glob

excel_files = [file[2:-5] for file in glob.glob('./*.xlsx')]
print(excel_files)
for file in excel_files:
    file_path = file
    wb = load_workbook(file_path + '.xlsx')
    xls = pd.ExcelFile(file_path + '.xlsx')
    sheets = wb.sheetnames

    dfs = []
    for sheet in sheets:
        df = pd.read_excel(xls, sheet)
        if "MFI" and "MFO" in list(df.head()):
            input_list_df = pd.DataFrame(
                [float(x.strip()[:-2]) * 1024 if x.strip()[-2:] == "Mb" else float(x.strip()[:-2]) for x in
                 list(df["Input"]) if str(x) != 'nan'])
            output_list_df = pd.DataFrame(
                [float(x.strip()[:-2]) * 1024 if x.strip()[-2:] == "Mb" else float(x.strip()[:-2]) for x in
                 list(df["Output"]) if str(x) != 'nan'])
            mfi_list_df = pd.DataFrame([x for x in list(df["MFI"]) if str(x) != 'nan'])
            mfo_list_df = pd.DataFrame([x for x in list(df["MFO"]) if str(x) != 'nan'])

            input_mf = input_list_df * mfi_list_df
            output_mf = output_list_df * mfo_list_df
            total = input_mf + output_mf
            try:
                df["Input(MF)"] = list(input_mf[0])
                df["Output(MF)"] = list(output_mf[0])
                df["Total"] = list(total[0])
            except ValueError:
                break
            else:
                dfs.append(df)
        else:
            dfs.append(pd.read_excel(xls, sheet))

            # # write to excel
            # df.to_excel(f"./DATA_SS/{file_path}_processed.xlsx", index=False)
            # writer = pd.ExcelWriter(f'./DATA_SS/{file_path}_processed.xlsx', engine='xlsxwriter')
            # df.to_excel(writer, sheet_name=sheet)
            # writer.save()
            # Create a Pandas Excel writer using XlsxWriter as the engine.

    Excelwriter = pd.ExcelWriter(f"./DATA_SS/{file_path}_processed.xlsx", engine="xlsxwriter")
    print(len(dfs))
    print(len(dfs) == 0)
    if len(dfs) == 0:
        continue
    else:
        for i, df in enumerate(dfs):
            df.to_excel(Excelwriter, sheet_name=sheets[i], index=False)

        Excelwriter.save()
